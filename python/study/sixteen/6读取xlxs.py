#!/usr/bin/python3.6
# -*- coding: utf-8 -*-
# @Author  : lzqian
# @Time    : 2020/10/9 14:32
# @File    : 6读取xlxs.py
# @Function:

from openpyxl.reader.excel import load_workbook
path = '1.xlsx'
file = load_workbook(filename=path)
table = file.get_sheet_names() # 获取所有表格的名称
# print(table)
# 拿出第一个表格
anli = file.get_sheet_by_name(table[0])
# print(anli.max_row) # 获取第一个表的当前行数
# print(anli.max_column) # 获取第一个表的当前列数

for num in range(1,anli.max_row+1):
    mylist =[]
    for col in range(1,anli.max_column+1):
        val = anli.cell(row=num,column=col).value # 根据行和列进行取值
        mylist.append(val) # 将数据添加到新的列表内
    print(mylist)