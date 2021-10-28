#!/usr/bin/python3.6
# -*- coding: utf-8 -*-
# @Author  : lzqian
# @Time    : 2020/10/9 15:01
# @File    : 7读取整个表格数据.py
# @Function:

from openpyxl.reader.excel import load_workbook

path = '1.xlsx'
file = load_workbook(path)
all_table = file.get_sheet_names()
for table in all_table:
    print(table)
    sheet = file.get_sheet_by_name(table) # 根据对当前的table进行操作

    for line in range(1,sheet.max_row+1):
        mylist = []
        for col in range(1,sheet.max_column+1):
            value = sheet.cell(row=line,column=col).value
            mylist.append(value)

        print(mylist)